from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys #點擊、輸入操作
from selenium.webdriver.support.ui import WebDriverWait #等待載入
from selenium.webdriver.support import expected_conditions as EC #等待載入
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import time     #輔助
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


def safe_load_workbook(file_path, retries=10, wait_seconds=1):
    """安全讀取 Excel，如果檔案被鎖住則重試。"""
    for attempt in range(retries):
        try:
            return load_workbook(file_path)
        except PermissionError:
            print(f"檔案正在使用中，嘗試重新載入（第 {attempt+1} 次）...")
            time.sleep(wait_seconds)
        except InvalidFileException:
            raise Exception("檔案格式錯誤，請確認是否為合法的 Excel 檔。")
    raise Exception("無法開啟 Excel 檔，請確認是否被其他程式鎖定。")

def safe_save_workbook(workbook, file_path, retries=10, wait_seconds=3):
    """安全儲存 Excel，如果檔案被鎖住則重試。"""
    for attempt in range(retries):
        try:
            workbook.save(file_path)
            return
        except PermissionError:
            print(f"無法儲存 Excel，檔案可能正被開啟中（第 {attempt+1} 次重試）...")
            time.sleep(wait_seconds)
    raise Exception("儲存失敗，請關閉 Excel 後再試。")

def load_config(config_path="config.json"):
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def setup_chrome_driver():
    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--disable-gpu')  # 禁用 GPU
    options.add_argument('--disable-software-rasterizer')
    options.add_argument('--no-sandbox')
    options.add_argument('--log-level=3')  # 僅顯示致命錯誤，關掉 Info/Warning
    options.add_experimental_option('excludeSwitches', ['enable-logging'])  # 關掉 DevTools 日誌
    
    driver = webdriver.Chrome(service=Service(), options=options)
    return driver


#警告處理
def handle_alert(driver):
    """處理網頁彈出的警告視窗。"""
    try:
        time.sleep(0.2)
        alert = driver.switch_to.alert
        print(alert.text)
        alert.accept()
        return 1
    except Exception:  # 更通用的異常處理
        pass
    return 0

def process_excel_data(file_path, search_col=1):
    """從 Excel 讀取資料，進行網頁查詢，並將結果寫回 Excel。"""
    driver = setup_chrome_driver()
    driver.get("https://serv.gcis.nat.gov.tw/Fidbweb/index.jsp")
    wait = WebDriverWait(driver, 10)

    search_frame = driver.find_element(By.NAME, 'search')
    driver.switch_to.frame(search_frame)

    try:
        wait.until(EC.visibility_of_element_located((By.NAME, 'regiID')))
    except Exception:
        raise TimeoutError('連線逾時，請關閉後重新操作')

    workbook = safe_load_workbook(file_path)
    worksheet = workbook.active

    # === 新增：在 E1~I1 寫入標題 ===
    result_headers = ["工廠位置", "工廠名稱", "工廠地址", "工廠編號", "營業狀況"]
    for index, header in enumerate(result_headers):
        col = get_column_letter(search_col + index + 1)  # search_col 為 D，+1 為 E 開始
        worksheet[f'{col}1'] = header

    row_count = 0

    for row_index in range(2, 10002):  # 從第二行開始，最多處理 10000 行
        row_count = row_index
        main_search_value = str(worksheet[f'{get_column_letter(search_col)}{row_index}'].value)

        if main_search_value == 'None':
            break

        print(f'Submiting {row_count-1} : {main_search_value}')

        search_input = driver.find_element(By.NAME, 'regiID')
        search_input.clear()
        search_input.send_keys(main_search_value)
        search_input.send_keys(Keys.RETURN)

        search_results = []
        if handle_alert(driver):
            search_results.append('資料無法查詢')
        else:
            driver.switch_to.parent_frame()
            result_frame = driver.find_element(By.NAME, 'show')
            driver.switch_to.frame(result_frame)
            search_results = perform_web_search(driver, wait)

        for index, res in enumerate(search_results):
            worksheet[f'{get_column_letter(search_col + index + 1)}{row_index}'] = res

        #workbook.save(file_path)
        safe_save_workbook(workbook, file_path)


    workbook.close()
    driver.quit()  # 關閉瀏覽器
    print('查詢結束，請至 Excel 確認結果')
    return row_count

def perform_web_search(driver,wait):
    """在網頁上進行搜尋，並擷取結果。"""
    try:
        wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/table/tbody/tr/td/table[2]/tbody/tr[1]/th[2]/font/h2')))
    except Exception:
        pass

    results = []
    result_link_xpath = '/html/body/form/table/tbody/tr/td/table[2]/tbody/tr[2]/td[2]/h3/a'
    try:
        result_link = driver.find_element(By.XPATH, result_link_xpath)
        result_link.click()
    except:
        results.append('查無資料')
        driver.switch_to.parent_frame()
        search_frame = driver.find_element(By.NAME, 'search')
        driver.switch_to.frame(search_frame)
        return results
        
    driver.switch_to.parent_frame()
    try:
        wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div/div[2]/div/div[3]/h2/b/font')))
    except Exception:
        raise TimeoutError('連線逾時')

    factory_name = driver.find_element(By.XPATH, '//*[@id="factInfoMain"]/b/font').text
    factory_number = driver.find_element(By.XPATH, '//*[@id="AutoNumber4"]/tbody/tr[1]/td[1]/font').text
    factory_address = driver.find_element(By.XPATH, '//*[@id="AutoNumber4"]/tbody/tr[2]/td/font').text
    factory_status = driver.find_element(By.XPATH, '//*[@id="AutoNumber4"]/tbody/tr[7]/td/font').text
    factory_location = driver.find_element(By.XPATH, '//*[@id="AutoNumber4"]/tbody/tr[3]/td[1]/font').text

    results.append(factory_location)
    results.append(factory_name)
    results.append(factory_address)
    results.append(factory_number)
    results.append(factory_status)

    driver.back()
    driver.switch_to.parent_frame()
    search_frame = driver.find_element(By.NAME, 'search')
    driver.switch_to.frame(search_frame)

    return results


if __name__ == '__main__':
    config = load_config()
    
    processed_rows = process_excel_data(config, 4)
    input(f'查詢了 {processed_rows - 2} 筆，任務完成')